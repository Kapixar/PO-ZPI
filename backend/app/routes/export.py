import datetime
from io import BytesIO
from flask import Blueprint, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.models.student import Student
from app.models.topic import TopicStatus


export_bp = Blueprint('export', __name__, url_prefix='/api/export')

@export_bp.route('/students-by-topic', methods=['GET'])
def export_students_by_topic():
    try:
        students = Student.query.order_by(
            Student.topic_id.nullslast(), 
            Student.index_number
        ).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Studenci wg tematów"
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = ["Topic ID", "Temat", "Nr indeksu", "Imię i Nazwisko", "Temat zatwierdzony przez KPK", "Prowadzący zatwierdził deklarację","Student zatwierdził deklarację"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        current_topic_id = None
        
        for student in students:
            if current_topic_id is not None and current_topic_id != student.topic_id:
                ws.append([])
            
            current_topic_id = student.topic_id
            
            row_data = [
                student.topic_id if student.topic_id else "Brak tematu",
                student.topic.title if student.topic else "Nie przypisany",
                student.index_number,
                student.account.full_name if student.account else "",
                "Tak" if student.topic and student.topic.status == TopicStatus.ZATWIERDZONY else "Nie",
                "Tak" if student.topic and student.topic.teacher and student.topic.teacher.is_declaration_approved else "Nie",
                "Tak" if student.is_declaration_approved else "Nie"
            ]

            ws.append(row_data)
        

        column_widths = [12, 60, 12, 30, 30, 30, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"studenci_tematy_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return {"error": str(e)}, 500